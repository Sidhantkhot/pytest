import openpyxl
from openpyxl import load_workbook


def get_row_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row


def get_column_count(file, sheetname):
    workbook = load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column


def read_xl_data(file, sheetname, rownum, col):
    workbook = load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum, column=col).value


def write_xl_data(file, sheetname, rownum, col, data):
    workbook = load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum, column=col).value = data
    workbook.save(file)
